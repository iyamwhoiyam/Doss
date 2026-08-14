#!/usr/bin/env python3
"""Refresh the compliance dashboard's embedded workbook snapshot.

Usage:  python3 refresh_from_workbook.py <tracker.xlsx> [<path/to/index.html>]

Reads every community sheet of the AH SWFL compliance tracker workbook and
rewrites the `const SEED = /*__SEED__*/...;` line inside the dashboard HTML.
One record per employee; item values are 'true'/'false'/ISO date/'N/A'/'-'
or the sheet's verbatim text."""
import openpyxl, datetime, json, re, sys, os

if len(sys.argv) < 2:
    sys.exit(__doc__)
SRC = sys.argv[1]
HTML = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "..", "index.html")

# canonical key -> list of header variants seen across sheets (normalized match)
ITEM_VARIANTS = {
    "app":      ["Employment Application"],
    "refs":     ["2 References"],
    "jd":       ["Job Description"],
    "ahca":     ["AHCA Acknowledgement & Screening"],
    "pform":    ["Personnel Form - New Hire"],
    "noauth":   ["Notice and Authorization"],
    "bg":       ["Criminal Background Check"],
    "offer":    ["Terms of Employment/Offer Letter"],
    "i9":       ["I-9 Verification"],
    "opack":    ["Orientation Packet (Signed)"],
    "ssn":      ["SSN Card"],
    "photoid":  ["Photo ID/Driver's License EXPIRATION DATE"],
    "passport": ["Passport EXPIRATION DATE"],
    "workauth": ["Employment Authorization EXPIRATION DATE"],
    "tb":       ["TB Test DATE (Update every 1 Yr)"],
    "xray":     ["Chest Xray DATE (Update Every 5 yrs)"],
    "commdis":  ["Communicable Disease Statement Form"],
    "drug":     ["Drug Test"],
    "cpr":      ["CPR/First Aid Cert EXPIRATION DATE (Update Every 2 Years)",
                 "CPR/First Aid Cert (Update Every 2 Years)"],
    "cnalpn":   ["CNA/LPN certs EXPIRATION DATE (Update Every 2 Years)",
                 "CNA/LPN certs (Update Every 2 Years)"],
    "abuse":    ["Abuse, Neglect & Exploit"],
    "alz1":     ["ALZ Level I", "FL ALZ & Dementia Level I"],
    "alz2":     ["ALZ Level II", "FL ALZ & Dementia Level II"],
    "adls":     ["Assistance w/ ADLS"],
    "biowaste": ["BioMedical Waste  (Update Annually)"],
    "dnr":      ["DNR Policy & Procedure"],
    "elope":    ["Elopement (Update Annually)"],
    "food":     ["Food Handling (Update Annually)"],
    "hipaa":    ["HIPAA"],
    "hiv":      ["HIV/AIDS (Every 2 Years)"],
    "traffick": ["Human Trafficking"],
    "infect":   ["Infection Control"],
    "emerg":    ["Major Emergency Planning & Preparedness",
                 "Major Emergencies, Planning, & Preparedness"],
    "preserv":  ["Pre-service Orientation"],
    "rights":   ["Resident Rights"],
    "sexhar":   ["Sexual Harassment in Long Term Care"],
    "sigchg":   ["Sig. changes & Adverse Incident Reporting"],
    "adrd":     ["ADRD 1 HOUR TRAINING."],
    "cares1":   ["CARES BASICS 5-STEP METHOD TRAINING"],
    "cares2":   ["CARES DEMENTIA RELATED BEHAVIORS TRAINING"],
    "flalz":    ["FL ALZ Update - 4 Hrs (Update Annually after 1st Year)"],
    "medtech":  ["6 Hour Med Tech Certification EXPIRATION DATE",
                 "6 Hour Med Tech Certification (Date of Expiration"],
    "selfmeds": ["Assistance w/ Self-Admin Meds 2 Hour Certification (Update Annually)"],
    "safestaff":["SafeStaff /ServFood (Every 3 Years)",
                 "SafeStaff /ServFood (Date of Expiration-Every 3 Years)"],
    "drill1":   ["Elopement Drill 1"],
    "drill2":   ["Elopement Drill 2"],
}
META = {"name": ["Employee Name + Number"], "hire": ["Hire Date"], "dept": ["Department-Position"]}

def norm(s):
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())

LOOKUP = {}
for k, variants in {**ITEM_VARIANTS, **META}.items():
    for v in variants:
        LOOKUP[norm(v)] = k

NA = {"n/a", "na", "n/a.", "n\\a"}
DASH = {"-", "- ", " -", "–", "—"}

def parse_loose_date(s):
    """Tolerant date parse for messy strings like '2/1//26', '03/06/024'."""
    m = re.search(r"(\d{1,2})\s*/+\s*(\d{1,2})\s*/+\s*(\d{2,4})", s)
    if not m:
        m2 = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s.strip())
        if m2:
            y, mo, d = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
            try: return datetime.date(y, mo, d)
            except ValueError: return None
        return None
    mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    elif y < 1000: y += 2000  # '024' -> 2024
    try: return datetime.date(y, mo, d)
    except ValueError: return None

def conv(v):
    if v is None: return ""
    if isinstance(v, bool): return "true" if v else "false"
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    s = str(v).strip()
    if not s: return ""
    low = s.lower()
    if low in ("true", "false"): return low
    if low in NA: return "N/A"
    if s in DASH: return "-"
    return s

def find_header_row(ws):
    for r in range(1, 6):
        if any(c.value and "Employee Name" in str(c.value) for c in ws[r]):
            return r
    raise RuntimeError(f"no header row in {ws.title}")

wb = openpyxl.load_workbook(SRC, data_only=True)
employees = []
unmapped = set()
next_id = 1
for ws in wb.worksheets:
    community = ws.title
    hr = find_header_row(ws)
    colmap = {}
    for i, c in enumerate(ws[hr], 1):
        if not c.value: continue
        key = LOOKUP.get(norm(c.value))
        if key: colmap[i] = key
        else: unmapped.add((community, str(c.value).strip()))
    name_col = [i for i, k in colmap.items() if k == "name"][0]
    for row in ws.iter_rows(min_row=hr + 1):
        cells = {i: c.value for i, c in enumerate(row, 1)}
        raw_name = cells.get(name_col)
        note = ""
        # Bonita/Coconut: col A carries a note or duplicates the name
        first = cells.get(1)
        if name_col > 1 and first is not None and str(first).strip():
            fs = str(first).strip()
            if not raw_name or not str(raw_name).strip():
                if " - 1" in fs:  # name landed in col A only
                    raw_name = fs
                else:
                    note = fs
            elif norm(fs) != norm(raw_name):
                note = fs
        if not raw_name or not str(raw_name).strip():
            continue
        name = re.sub(r"\s+", " ", str(raw_name).strip())
        rec = {"id": f"e{next_id}", "name": name, "community": community,
               "dept": "", "hire": "", "note": note, "items": {}}
        next_id += 1
        for i, key in colmap.items():
            if key == "name": continue
            v = conv(cells.get(i))
            if key == "dept":
                rec["dept"] = v
            elif key == "hire":
                if v:
                    d = parse_loose_date(v) if not re.match(r"^\d{4}-\d{2}-\d{2}$", v) else None
                    rec["hire"] = d.isoformat() if d else (v if re.match(r"^\d{4}-\d{2}-\d{2}$", v) else "")
                    if not rec["hire"] and v:
                        print(f"  ! unparseable hire date {v!r} for {name} ({community})")
            else:
                if v != "":
                    rec["items"][key] = v
        employees.append(rec)

print("employees:", len(employees))
from collections import Counter
print(Counter(e["community"] for e in employees))
if unmapped:
    print("UNMAPPED HEADERS:", unmapped)
notes = [(e["name"], e["note"]) for e in employees if e["note"]]
print("notes captured:", len(notes), notes[:8])
blank_hire = sum(1 for e in employees if not e["hire"])
print("blank hire dates:", blank_hire)
depts = Counter(e["dept"].split("-")[0].strip() for e in employees if e["dept"])
print("dept prefixes:", dict(depts))

seed = json.dumps(employees, separators=(",", ":"))
html = open(HTML).read()
new_html, n = re.subn(r"const SEED = /\*__SEED__\*/.*?;\n",
                      "const SEED = /*__SEED__*/" + seed.replace("\\", "\\\\") + ";\n",
                      html, count=1, flags=re.S)
if n != 1:
    sys.exit("ERROR: SEED marker not found in " + HTML)
open(HTML, "w").write(new_html)
print("updated", HTML, "with", len(employees), "employees (", len(seed), "bytes of data )")
